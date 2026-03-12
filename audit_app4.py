"""
АУДИТЫН ХОУ ПРОТОТИП v5.0 — САЙЖРУУЛСАН
═══════════════════════════════════════════
Шинэ боломжууд:
  1️⃣ Universal Parser — SAP, 1C, QuickBooks, ямар ч ERP
  2️⃣ Audit Risk Rules Engine — ISA-д суурилсан дүрмүүд
  3️⃣ ML Upgrade — Autoencoder, LOF, DBSCAN, Weighted Ensemble
  4️⃣ Dashboard Upgrade — Executive Summary, Drill-down, Audit Trail
  5️⃣ Performance — 3M+ мөр, chunked processing, caching

pip install streamlit pandas numpy scikit-learn plotly openpyxl
streamlit run audit_app5.py
"""
import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from sklearn.ensemble import (
    IsolationForest, RandomForestClassifier, 
    GradientBoostingClassifier, VotingClassifier
)
from sklearn.linear_model import LogisticRegression
from sklearn.neighbors import LocalOutlierFactor
from sklearn.cluster import DBSCAN
from sklearn.neural_network import MLPRegressor
from sklearn.preprocessing import LabelEncoder, StandardScaler, MinMaxScaler
from sklearn.model_selection import StratifiedKFold, cross_val_predict
from sklearn.metrics import (
    precision_score, recall_score, f1_score, 
    roc_auc_score, roc_curve, confusion_matrix
)
import warnings, io, re, gzip, hashlib, time, json
from datetime import datetime, timedelta
from collections import Counter
from functools import lru_cache
warnings.filterwarnings('ignore')

# ═══════════════════════════════════════
# ТОХИРГОО
# ═══════════════════════════════════════
APP_VERSION = "5.0"
MAX_TXN_FULL = 500_000      # Гүйлгээний бүрэн шинжилгээний дээд хязгаар
CHUNK_SIZE = 100_000         # Chunk бүрийн хэмжээ
SAMPLE_SIZE_DEFAULT = 100_000 # Анхдагч sampling хэмжээ

st.set_page_config(page_title=f"Аудитын ХОУ v{APP_VERSION}", page_icon="🔍", layout="wide")

# ═══════════════════════════════════════
# 🎨 CUSTOM CSS
# ═══════════════════════════════════════
st.markdown("""
<style>
    .main-header { text-align: center; color: #1565c0; font-size: 2em; font-weight: bold; margin-bottom: 5px; }
    .sub-header { text-align: center; color: #666; font-size: 0.9em; margin-bottom: 20px; }
    .metric-card { background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
                   padding: 15px; border-radius: 12px; color: white; text-align: center; }
    .risk-high { background-color: #ffebee; border-left: 4px solid #c62828; padding: 10px; border-radius: 8px; }
    .risk-med { background-color: #fff3e0; border-left: 4px solid #ef6c00; padding: 10px; border-radius: 8px; }
    .risk-low { background-color: #e8f5e9; border-left: 4px solid #2e7d32; padding: 10px; border-radius: 8px; }
    .info-box { background-color: #e3f2fd; padding: 15px; border-radius: 8px; border-left: 4px solid #1565c0; margin-bottom: 15px; }
    .stTabs [data-baseweb="tab-list"] { gap: 2px; }
    .stTabs [data-baseweb="tab"] { padding: 8px 16px; }
</style>
""", unsafe_allow_html=True)

st.markdown(f'<div class="main-header">🔍 Аудитын хиймэл оюуны прототип v{APP_VERSION}</div>', unsafe_allow_html=True)
st.markdown('<div class="sub-header">Universal Parser • Risk Rules Engine • ML Ensemble • 3M+ Performance</div>', unsafe_allow_html=True)

# ═══════════════════════════════════════
# 1️⃣ UNIVERSAL PARSER
# ═══════════════════════════════════════

# Expanded column patterns for multiple ERP systems
COL_PATTERNS = {
    'account_code': ['дансны код','данс код','account code','account no','account number','acc code','код',
                     'gl account','кт','sap account','счет','номер счета','acct','account_code',
                     'chart of accounts','coa code','ledger account'],
    'account_name': ['дансны нэр','данс нэр','account name','acc name','нэр','наименование счета',
                     'account description','gl description','account_name'],
    'transaction_date': ['огноо','date','transaction date','txn date','posting date','document date',
                         'дата','дата проводки','entry date','value date','effective date'],
    'debit_mnt': ['дебит','debit','dt','дт','debit amount','debit_mnt','сумма дебет','debit amt',
                  'debit_amount','dr amount','dr','debit (mnt)'],
    'credit_mnt': ['кредит','credit','ct','кт','credit amount','credit_mnt','сумма кредит','credit amt',
                   'credit_amount','cr amount','cr','credit (mnt)'],
    'balance_mnt': ['үлдэгдэл','balance','bal','ending balance','closing balance','сальдо','остаток',
                    'running balance','balance_mnt'],
    'counterparty_name': ['харилцагч','counterparty','partner','vendor','customer','контрагент',
                          'trading partner','bp name','business partner','supplier','customer name'],
    'transaction_description': ['тайлбар','гүйлгээний утга','утга','description','memo','narration',
                                'назначение','содержание','text','line description','reference',
                                'transaction text','posting text'],
    'journal_no': ['журнал','journal','journal no','journal number','journal entry','номер журнала',
                   'je number','entry number','batch','batch no'],
    'document_no': ['баримт','document','doc no','document number','invoice','номер документа',
                    'reference no','ref no','voucher','voucher no'],
    'currency': ['валют','currency','curr','ccy','валюта'],
    'amount': ['дүн','amount','total','сумма','нийт дүн','net amount'],
    'cost_center': ['зардлын төв','cost center','cc','центр затрат','department','dept'],
    'project': ['төсөл','project','проект','project code'],
}

# ERP-specific format signatures
ERP_SIGNATURES = {
    'sap': {
        'columns': ['posting date','document number','reference','debit/credit ind','amount in local cur'],
        'patterns': [r'FI\d{8}', r'\d{10}'],  # SAP document numbers
    },
    '1c': {
        'columns': ['дата','номер','сумма','дебет','кредит','содержание'],
        'patterns': [r'БП-\d+', r'ПКО-\d+'],
    },
    'quickbooks': {
        'columns': ['txn type','date','name','memo','account','debit','credit'],
        'patterns': [r'(Invoice|Bill|Check|Journal)'],
    },
    'oracle': {
        'columns': ['journal name','line number','entered debit','entered credit','accounting date'],
        'patterns': [r'JE-\d+'],
    },
    'mongolian_edt': {
        'columns': [],
        'patterns': [r'Данс:\s*\[', r'Компани:', r'ЕРӨНХИЙ'],
    },
    'mongolian_tb': {
        'columns': [],
        'patterns': [r'\d{3}-\d{2}-\d{2}-\d{3}'],
    },
}

def detect_encoding(raw_bytes):
    """Detect file encoding (UTF-8, CP1251, GB2312, etc.)"""
    encodings = ['utf-8', 'utf-8-sig', 'cp1251', 'cp949', 'gb2312', 'latin-1', 'iso-8859-1']
    for enc in encodings:
        try:
            raw_bytes.decode(enc)
            return enc
        except (UnicodeDecodeError, AttributeError):
            continue
    return 'utf-8'

def fuzzy_match_column(header, field, threshold=0.6):
    """Fuzzy match column header to field name."""
    h = str(header).lower().strip()
    patterns = COL_PATTERNS.get(field, [])
    # Exact substring match
    for p in patterns:
        if p in h or h in p:
            return True
    # Partial word match
    h_words = set(h.split())
    for p in patterns:
        p_words = set(p.split())
        if len(h_words & p_words) > 0 and len(h_words & p_words) / max(len(p_words), 1) >= threshold:
            return True
    return False

def auto_map_columns(headers):
    """Enhanced column mapping with fuzzy matching."""
    mapping, used = {}, set()
    priority_fields = [
        'account_code', 'debit_mnt', 'credit_mnt', 'transaction_date',
        'account_name', 'counterparty_name', 'transaction_description',
        'balance_mnt', 'journal_no', 'document_no', 'amount', 'currency',
        'cost_center', 'project'
    ]
    for field in priority_fields:
        for i, h in enumerate(headers):
            if i in used:
                continue
            if fuzzy_match_column(h, field):
                mapping[field] = i
                used.add(i)
                break
    # If only 'amount' found (no separate debit/credit), handle single amount column
    if 'amount' in mapping and 'debit_mnt' not in mapping and 'credit_mnt' not in mapping:
        mapping['single_amount'] = mapping.pop('amount')
    return mapping

def detect_erp_type(headers, sample_rows):
    """Detect ERP system type from headers and data patterns."""
    headers_lower = [str(h).lower().strip() for h in headers if h]
    all_text = ' '.join(headers_lower)
    
    for erp_name, sig in ERP_SIGNATURES.items():
        # Check column patterns
        col_matches = sum(1 for col in sig['columns'] if any(col in h for h in headers_lower))
        if col_matches >= 2:
            return erp_name
        
        # Check data patterns
        for row in sample_rows[:20]:
            row_text = ' '.join(str(c) for c in row if c)
            for pat in sig['patterns']:
                if re.search(pat, row_text):
                    return erp_name
    return 'generic'

# ═══════════════════════════════════════
# 2️⃣ AUDIT RISK RULES ENGINE
# ═══════════════════════════════════════

class AuditRiskEngine:
    """ISA-д суурилсан аудитын эрсдэлийн дүрмүүд."""
    
    # Монголын баярын өдрүүд
    MN_HOLIDAYS = [
        (1, 1), (1, 2),   # Шинэ жил
        (3, 8),            # Олон улсын эмэгтэйчүүдийн өдөр
        (6, 1),            # Хүүхдийн өдөр
        (7, 11), (7, 12), (7, 13),  # Наадам
        (11, 26),          # Тусгаар тогтнолын өдөр
        (12, 29),          # Үндэсний эрх чөлөөний өдөр
    ]
    
    @staticmethod
    def check_weekend_holiday(df):
        """Амралтын өдөр, баярын гүйлгээ илрүүлэх (ISA 240)."""
        if 'transaction_date' not in df.columns:
            df['is_weekend'] = 0
            df['is_holiday'] = 0
            return df
        
        dates = pd.to_datetime(df['transaction_date'], errors='coerce')
        df['is_weekend'] = dates.dt.dayofweek.isin([5, 6]).fillna(False).astype(int)
        
        # Holiday check
        month_day = list(zip(dates.dt.month.fillna(0).astype(int), dates.dt.day.fillna(0).astype(int)))
        df['is_holiday'] = [1 if (m, d) in AuditRiskEngine.MN_HOLIDAYS else 0 for m, d in month_day]
        
        return df
    
    @staticmethod
    def check_round_amounts(df, thresholds=None):
        """Бүхэл тэгс дүнтэй гүйлгээ илрүүлэх (ISA 240)."""
        if thresholds is None:
            thresholds = [1_000_000, 10_000_000, 100_000_000]
        
        amt = df['amount'] if 'amount' in df.columns else (df.get('debit_mnt', 0).abs() + df.get('credit_mnt', 0).abs())
        
        df['is_round_1m'] = ((amt >= 1e6) & (amt % 1e6 == 0)).astype(int)
        df['is_round_10m'] = ((amt >= 1e7) & (amt % 1e7 == 0)).astype(int)
        df['is_round_100m'] = ((amt >= 1e8) & (amt % 1e8 == 0)).astype(int)
        df['round_score'] = df['is_round_1m'] + df['is_round_10m'] * 2 + df['is_round_100m'] * 3
        
        return df
    
    @staticmethod
    def check_just_below_threshold(df, materiality=None):
        """Материаллагийн босгоос яг доогуур дүнтэй гүйлгээ (ISA 320)."""
        amt = df['amount'] if 'amount' in df.columns else (df.get('debit_mnt', 0).abs() + df.get('credit_mnt', 0).abs())
        
        if materiality is None:
            # Auto-calculate materiality as 1% of total turnover
            materiality = amt.sum() * 0.01
        
        # Check if amount is within 90-99% of materiality
        df['near_materiality'] = ((amt >= materiality * 0.90) & (amt < materiality)).astype(int)
        
        return df
    
    @staticmethod
    def check_benford_per_account(df):
        """Данс тус бүрийн Бенфордын хуулийн шинжилгээ (ISA 240)."""
        benford_expected = {1: 0.301, 2: 0.176, 3: 0.125, 4: 0.097, 5: 0.079, 6: 0.067, 7: 0.058, 8: 0.051, 9: 0.046}
        
        amt = df['amount'] if 'amount' in df.columns else (df.get('debit_mnt', 0).abs() + df.get('credit_mnt', 0).abs())
        first_digits = amt.apply(lambda x: int(str(int(abs(x)))[0]) if abs(x) >= 1 else 0)
        
        benford_scores = {}
        for acct in df['account_code'].unique():
            mask = (df['account_code'] == acct) & (first_digits > 0)
            if mask.sum() < 30:  # Need minimum 30 transactions for Benford
                continue
            observed = first_digits[mask].value_counts(normalize=True)
            chi2 = sum(((observed.get(d, 0) - benford_expected.get(d, 0)) ** 2) / benford_expected.get(d, 0.001)
                      for d in range(1, 10))
            benford_scores[acct] = chi2
        
        # Flag accounts with chi2 > 15.51 (critical value at p=0.05, df=8)
        df['benford_acct_chi2'] = df['account_code'].map(benford_scores).fillna(0)
        df['benford_acct_flag'] = (df['benford_acct_chi2'] > 15.51).astype(int)
        
        return df
    
    @staticmethod
    def check_segregation_of_duties(df):
        """Нэг хүн олон дансанд хийсэн гүйлгээ (ISA 315)."""
        if 'counterparty_name' not in df.columns:
            df['sod_risk'] = 0
            return df
        
        # Check if same counterparty appears in both asset and liability accounts
        cp_accounts = df.groupby('counterparty_name')['account_code'].apply(
            lambda x: set(s[0] for s in x.astype(str) if len(s) > 0)
        )
        
        sod_risky = {}
        for cp, acct_types in cp_accounts.items():
            # If same counterparty has transactions in both income and expense
            if ('5' in acct_types or '6' in acct_types) and ('7' in acct_types or '8' in acct_types):
                sod_risky[cp] = 1
            # If same counterparty in both asset and liability
            elif '1' in acct_types and '2' in acct_types:
                sod_risky[cp] = 1
            else:
                sod_risky[cp] = 0
        
        df['sod_risk'] = df['counterparty_name'].map(sod_risky).fillna(0).astype(int)
        return df
    
    @staticmethod
    def check_sequential_entries(df):
        """Дараалсан дугаартай гүйлгээнүүдийн зөрчил (ISA 500)."""
        if 'document_no' not in df.columns or 'transaction_date' not in df.columns:
            df['seq_gap'] = 0
            return df
        
        # Check for gaps in document numbering
        try:
            doc_nums = pd.to_numeric(df['document_no'], errors='coerce').dropna()
            if len(doc_nums) > 10:
                sorted_nums = doc_nums.sort_values()
                gaps = sorted_nums.diff()
                gap_threshold = gaps.median() * 5  # 5x median gap = suspicious
                df['seq_gap'] = 0
                df.loc[doc_nums.index, 'seq_gap'] = (gaps > gap_threshold).fillna(0).astype(int)
            else:
                df['seq_gap'] = 0
        except:
            df['seq_gap'] = 0
        
        return df
    
    @classmethod
    def apply_all_rules(cls, df, progress_callback=None):
        """Бүх дүрмүүдийг нэг дор ажиллуулах."""
        steps = [
            ("Амралтын өдөр шалгах", cls.check_weekend_holiday),
            ("Тэгс дүн шалгах", cls.check_round_amounts),
            ("Материаллаг босго шалгах", cls.check_just_below_threshold),
            ("Бенфорд (данс тус бүр)", cls.check_benford_per_account),
            ("Үүрэг хуваарилалт", cls.check_segregation_of_duties),
            ("Дараалал шалгах", cls.check_sequential_entries),
        ]
        
        for i, (name, func) in enumerate(steps):
            try:
                df = func(df)
            except Exception as e:
                pass  # Дүрэм алдаа гарвал алгасна
            if progress_callback:
                progress_callback((i + 1) / len(steps), f"🔍 {name}...")
        
        return df


# ═══════════════════════════════════════
# 3️⃣ ML MODEL UPGRADE
# ═══════════════════════════════════════

class MLEngine:
    """Сайжруулсан ML загваруудын нэгтгэл."""
    
    @staticmethod
    def autoencoder_anomaly(X, contamination=0.1):
        """Autoencoder-д суурилсан аномали илрүүлэлт."""
        scaler = MinMaxScaler()
        X_scaled = scaler.fit_transform(X)
        
        # Train autoencoder (encoder-decoder)
        n_features = X_scaled.shape[1]
        hidden_size = max(n_features // 2, 3)
        
        ae = MLPRegressor(
            hidden_layer_sizes=(hidden_size, hidden_size // 2, hidden_size),
            activation='relu',
            max_iter=200,
            random_state=42,
            early_stopping=True,
            validation_fraction=0.1,
        )
        ae.fit(X_scaled, X_scaled)
        
        # Reconstruction error
        X_pred = ae.predict(X_scaled)
        mse = np.mean((X_scaled - X_pred) ** 2, axis=1)
        
        # Flag top contamination% as anomalies
        threshold = np.percentile(mse, (1 - contamination) * 100)
        anomalies = (mse > threshold).astype(int)
        
        return anomalies, mse
    
    @staticmethod
    def lof_anomaly(X, contamination=0.1):
        """Local Outlier Factor аномали илрүүлэлт."""
        lof = LocalOutlierFactor(
            n_neighbors=20,
            contamination=contamination,
            novelty=False,
            n_jobs=-1
        )
        labels = lof.fit_predict(X)
        scores = -lof.negative_outlier_factor_
        return (labels == -1).astype(int), scores
    
    @staticmethod
    def dbscan_anomaly(X, eps=0.5, min_samples=5):
        """DBSCAN кластерийн аномали илрүүлэлт."""
        scaler = StandardScaler()
        X_scaled = scaler.fit_transform(X)
        
        db = DBSCAN(eps=eps, min_samples=min_samples, n_jobs=-1)
        labels = db.fit_predict(X_scaled)
        
        # Noise points (label=-1) are anomalies
        return (labels == -1).astype(int), labels
    
    @staticmethod
    def weighted_ensemble(predictions, weights=None):
        """Жинлэсэн санал нэгтгэл."""
        if weights is None:
            weights = [1.0] * len(predictions)
        
        total_weight = sum(weights)
        weighted_sum = sum(p * w for p, w in zip(predictions, weights))
        
        # Threshold: weighted average > 0.5
        return (weighted_sum / total_weight >= 0.5).astype(int)


# ═══════════════════════════════════════
# 5️⃣ PERFORMANCE — Cached functions
# ═══════════════════════════════════════

@st.cache_data(ttl=3600, show_spinner=False)
def cached_read_excel(file_bytes, sheet_name=None):
    """Cached Excel reading."""
    return pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name)

@st.cache_data(ttl=3600, show_spinner=False)
def cached_read_csv(file_bytes, dtype=None):
    """Cached CSV reading."""
    raw = file_bytes
    if raw[:2] == b'\x1f\x8b':
        raw = gzip.decompress(raw)
    encoding = detect_encoding(raw)
    return pd.read_csv(io.BytesIO(raw if isinstance(raw, bytes) else raw.encode()), dtype=dtype, encoding=encoding)

def process_in_chunks(df, func, chunk_size=CHUNK_SIZE, progress_bar=None):
    """Process large DataFrames in chunks with progress."""
    if len(df) <= chunk_size:
        return func(df)
    
    results = []
    n_chunks = (len(df) + chunk_size - 1) // chunk_size
    
    for i in range(n_chunks):
        start = i * chunk_size
        end = min((i + 1) * chunk_size, len(df))
        chunk = df.iloc[start:end].copy()
        results.append(func(chunk))
        
        if progress_bar:
            progress_bar.progress((i + 1) / n_chunks)
    
    return pd.concat(results, ignore_index=True)

def optimize_dtypes(df):
    """Санах ойн хэрэглээг бууруулах dtype optimization."""
    for col in df.columns:
        if df[col].dtype == 'float64':
            if df[col].abs().max() < 1e10:
                df[col] = df[col].astype('float32')
        elif df[col].dtype == 'int64':
            if df[col].max() < 2**31:
                df[col] = df[col].astype('int32')
        elif df[col].dtype == 'object':
            if df[col].nunique() / len(df) < 0.5:
                df[col] = df[col].astype('category')
    return df


# ═══════════════════════════════════════
# ORIGINAL FUNCTIONS (сайжруулсан)
# ═══════════════════════════════════════

ACCT_RE_B = re.compile(r'Данс:\s*\[([^\]]+)\]\s*(.*)')
ACCT_RE_P = re.compile(r'Данс:\s*(\d{3}-\d{2}-\d{2}-\d{3})\s+(.*)')

def parse_account(text):
    m = ACCT_RE_B.match(text)
    if m:
        return m.group(1).strip(), m.group(2).strip()
    m = ACCT_RE_P.match(text)
    if m:
        return m.group(1).strip(), m.group(2).strip()
    return None, None

def safe_float(v):
    if v is None or v == '':
        return 0.0
    try:
        return float(v)
    except Exception:
        return 0.0

def process_raw_tb(file_obj):
    """ГҮЙЛГЭЭ_БАЛАНС → TB_standardized хөрвүүлэлт."""
    import openpyxl
    wb = openpyxl.load_workbook(file_obj, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = []
    for row in ws.iter_rows(values_only=True):
        if row[0] is None:
            continue
        try:
            int(float(row[0]))
        except Exception:
            continue
        code = str(row[1]).strip() if row[1] else ''
        if not code or not re.match(r'\d{3}-', code):
            continue
        rows.append({
            'account_code': code,
            'account_name': str(row[2]).strip() if row[2] else '',
            'opening_debit': safe_float(row[3]),
            'opening_credit': safe_float(row[4]),
            'turnover_debit': safe_float(row[5]),
            'turnover_credit': safe_float(row[6]),
            'closing_debit': safe_float(row[7]),
            'closing_credit': safe_float(row[8]),
        })
    wb.close()
    df = pd.DataFrame(rows)
    df['opening_balance_signed'] = df['opening_debit'] - df['opening_credit']
    df['turnover_net_signed'] = df['turnover_debit'] - df['turnover_credit']
    df['closing_balance_signed'] = df['closing_debit'] - df['closing_credit']
    df['net_change_signed'] = df['closing_balance_signed'] - df['opening_balance_signed']
    tb_sum = df[['account_code','account_name','opening_debit','opening_credit','opening_balance_signed',
                  'turnover_debit','turnover_credit','turnover_net_signed',
                  'closing_debit','closing_credit','closing_balance_signed','net_change_signed']].copy()
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as w:
        df[['account_code','account_name','opening_debit','opening_credit','turnover_debit','turnover_credit','closing_debit','closing_credit']].to_excel(w, sheet_name='01_TB_CLEAN', index=False)
        tb_sum.to_excel(w, sheet_name='02_ACCOUNT_SUMMARY', index=False)
    buf.seek(0)
    return buf, tb_sum


def process_edt(file_obj, report_year):
    """Universal ЕДТ/Ерөнхий журнал уншигч — олон ERP дэмждэг."""
    import openpyxl
    EDT_COLUMNS = ['report_year','account_code','account_name','transaction_no','transaction_date',
                   'journal_no','document_no','counterparty_name','counterparty_id',
                   'transaction_description','debit_mnt','credit_mnt','balance_mnt','month']

    # 1. Монголын стандарт ЕДТ формат
    file_obj.seek(0)
    wb = openpyxl.load_workbook(file_obj, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows_out, cur_code, cur_name = [], None, None
    for row in ws.iter_rows(values_only=True):
        c0 = row[0]
        if c0 is None: continue
        s = str(c0).strip()
        if s.startswith('Данс:'):
            code, name = parse_account(s)
            if code: cur_code, cur_name = code, name
            continue
        if any(s.startswith(x) for x in ['Компани:','ЕРӨНХИЙ','Тайлант','Үүсгэсэн','Журнал:','№','Эцсийн','Дт -','Нийт','Эхний','Нээгээд']) or s in ('Валютаар','Төгрөгөөр',''): continue
        try: tx_no = int(float(c0))
        except: continue
        if cur_code is None: continue
        td_val = row[1] if len(row)>1 else ''
        tx_date = td_val.strftime('%Y-%m-%d') if isinstance(td_val, datetime) else (str(td_val).strip() if td_val else '')
        rows_out.append({'report_year':str(report_year),'account_code':cur_code,'account_name':cur_name,
            'transaction_no':str(tx_no),'transaction_date':tx_date,
            'journal_no':str(row[5]).strip() if len(row)>5 and row[5] else '',
            'document_no':str(row[6]).strip() if len(row)>6 and row[6] else '',
            'counterparty_name':str(row[3]).strip() if len(row)>3 and row[3] else '',
            'counterparty_id':str(row[4]).strip() if len(row)>4 and row[4] else '',
            'transaction_description':str(row[7]).strip() if len(row)>7 and row[7] else '',
            'debit_mnt':safe_float(row[9]) if len(row)>9 else 0.0,
            'credit_mnt':safe_float(row[11]) if len(row)>11 else 0.0,
            'balance_mnt':safe_float(row[13]) if len(row)>13 else 0.0,
            'month':tx_date[:7] if len(tx_date)>=7 else ''})
    wb.close()
    if rows_out:
        return pd.DataFrame(rows_out), len(rows_out)

    # 2. Хүснэгт формат (SAP, Oracle, 1C, QuickBooks, generic)
    file_obj.seek(0)
    try:
        raw = file_obj.read(); file_obj.seek(0)
        wb2 = openpyxl.load_workbook(io.BytesIO(raw), read_only=True)
        ws2 = wb2[wb2.sheetnames[0]]
        allr = []
        for i, row in enumerate(ws2.iter_rows(values_only=True)):
            allr.append(list(row))
            if i >= 1000: break  # Increased from 500 to 1000 for better detection
        wb2.close()
        
        # Find header row
        best_i, best_s = 0, 0
        for i, row in enumerate(allr[:30]):  # Check first 30 rows
            sc = sum(1 for cell in row if cell and any(
                fuzzy_match_column(cell, f) for f in COL_PATTERNS.keys()
            ))
            if sc > best_s: best_s = sc; best_i = i
        
        if best_s >= 2:
            headers = [str(c).strip() if c else f'col_{j}' for j, c in enumerate(allr[best_i])]
            
            # Detect ERP type
            erp_type = detect_erp_type(headers, allr[best_i+1:best_i+20])
            
            cm = auto_map_columns(headers)
            
            # Handle single amount column (SAP-style)
            has_single_amount = 'single_amount' in cm
            
            if 'debit_mnt' in cm or 'credit_mnt' in cm or has_single_amount:
                def _gv(row, cm, f, d=''):
                    idx = cm.get(f)
                    if idx is None or idx >= len(row) or row[idx] is None: return d
                    return str(row[idx]).strip()
                
                rows2 = []
                for row in allr[best_i+1:]:
                    if all(c is None for c in row): continue
                    ac = _gv(row, cm, 'account_code')
                    if not ac or ac in ('None','nan',''): continue
                    
                    if has_single_amount:
                        amt = safe_float(row[cm['single_amount']])
                        db = max(amt, 0)
                        cr = abs(min(amt, 0))
                    else:
                        db = safe_float(row[cm['debit_mnt']]) if 'debit_mnt' in cm and cm['debit_mnt'] < len(row) else 0.0
                        cr = safe_float(row[cm['credit_mnt']]) if 'credit_mnt' in cm and cm['credit_mnt'] < len(row) else 0.0
                    
                    if db == 0 and cr == 0: continue
                    
                    tdi = cm.get('transaction_date')
                    tx_date = ''
                    if tdi is not None and tdi < len(row):
                        td2 = row[tdi]
                        tx_date = td2.strftime('%Y-%m-%d') if isinstance(td2, datetime) else (str(td2).strip()[:10] if td2 else '')
                    
                    rows2.append({
                        'report_year': str(report_year),
                        'account_code': ac,
                        'account_name': _gv(row, cm, 'account_name'),
                        'transaction_no': str(len(rows2) + 1),
                        'transaction_date': tx_date,
                        'journal_no': _gv(row, cm, 'journal_no'),
                        'document_no': _gv(row, cm, 'document_no'),
                        'counterparty_name': _gv(row, cm, 'counterparty_name'),
                        'counterparty_id': '',
                        'transaction_description': _gv(row, cm, 'transaction_description'),
                        'debit_mnt': db,
                        'credit_mnt': cr,
                        'balance_mnt': safe_float(row[cm['balance_mnt']]) if 'balance_mnt' in cm and cm['balance_mnt'] < len(row) else 0.0,
                        'month': tx_date[:7] if len(tx_date) >= 7 else ''
                    })
                if rows2:
                    return pd.DataFrame(rows2), len(rows2)
    except:
        pass

    # 3. pandas-аар шууд
    file_obj.seek(0)
    try:
        df = pd.read_excel(file_obj)
        cm = auto_map_columns(df.columns.tolist())
        if 'debit_mnt' in cm or 'credit_mnt' in cm:
            rn = {df.columns[idx]: field for field, idx in cm.items()}
            df = df.rename(columns=rn)
            df['report_year'] = str(report_year)
            for c in EDT_COLUMNS:
                if c not in df.columns:
                    df[c] = '' if c in ('account_code','account_name','transaction_description','counterparty_name') else 0
            df['debit_mnt'] = pd.to_numeric(df.get('debit_mnt', 0), errors='coerce').fillna(0)
            df['credit_mnt'] = pd.to_numeric(df.get('credit_mnt', 0), errors='coerce').fillna(0)
            df = df[(df['debit_mnt'] != 0) | (df['credit_mnt'] != 0)]
            df['month'] = df['transaction_date'].astype(str).str[:7] if 'transaction_date' in df.columns else ''
            if len(df) > 0:
                return df[EDT_COLUMNS], len(df)
    except:
        pass

    return pd.DataFrame(columns=EDT_COLUMNS), 0


# ═══════════════════════════════════════
# ШИНЖ ЧАНАР ҮҮСГЭХ (сайжруулсан)
# ═══════════════════════════════════════

def engineer_txn_features_v5(d, progress_callback=None):
    """v5.0 — Сайжруулсан шинж чанар үүсгэх: 20+ features."""
    d = d.copy()
    
    # Баганууд
    for c in ['debit_mnt','credit_mnt','account_code','account_name','counterparty_name','transaction_description','transaction_date']:
        if c not in d.columns:
            d[c] = '' if c in ('account_code','account_name','counterparty_name','transaction_description','transaction_date') else 0
    
    d['debit_mnt'] = pd.to_numeric(d['debit_mnt'], errors='coerce').fillna(0)
    d['credit_mnt'] = pd.to_numeric(d['credit_mnt'], errors='coerce').fillna(0)
    d['account_code'] = d['account_code'].astype(str).fillna('000')
    d['account_name'] = d['account_name'].astype(str).fillna('')
    d['counterparty_name'] = d['counterparty_name'].astype(str).fillna('')
    d['transaction_description'] = d['transaction_description'].astype(str).fillna('')
    d['transaction_date'] = d['transaction_date'].astype(str).fillna('')

    d['amount'] = d['debit_mnt'].abs() + d['credit_mnt'].abs()
    d['log_amount'] = np.log1p(d['amount'])
    d['is_debit'] = (d['debit_mnt'] > 0).astype(int)

    if progress_callback: progress_callback(0.1, "Дансны ангилал...")
    
    # Дансны ангилал
    try:
        le2 = LabelEncoder()
        d['acct_cat_num'] = le2.fit_transform(d['account_code'].str[:3])
    except:
        d['acct_cat_num'] = 0

    # Бенфорд
    digits = d['amount'].apply(lambda x: int(str(int(abs(x)))[0]) if abs(x) >= 1 else 0)
    d['benford_digit'] = digits
    benford_exp = {1:0.301,2:0.176,3:0.125,4:0.097,5:0.079,6:0.067,7:0.058,8:0.051,9:0.046}
    af = d[d['benford_digit'] > 0]['benford_digit'].value_counts(normalize=True)
    d['benford_dev'] = d['benford_digit'].map(lambda x: abs(af.get(x, 0) - benford_exp.get(x, 0)) if x > 0 else 0)

    if progress_callback: progress_callback(0.3, "Тэгс тоо, Z-score...")
    
    # Тэгс тоо
    d['is_round'] = (((d['amount'] >= 1e6) & (d['amount'] % 1e6 == 0)).astype(int) + 
                     ((d['amount'] >= 1e3) & (d['amount'] % 1e3 == 0)).astype(int))

    # Z-score
    try:
        as2 = d.groupby('account_code')['amount'].agg(['mean','std']).fillna(0)
        as2.columns = ['acct_mean','acct_std']
        d = d.merge(as2, on='account_code', how='left')
        d['amt_zscore'] = np.where(d['acct_std'] > 0, (d['amount'] - d['acct_mean']) / d['acct_std'], 0)
        d['amt_zscore'] = d['amt_zscore'].clip(-10, 10).fillna(0)
    except:
        d['acct_mean'] = 0; d['acct_std'] = 0; d['amt_zscore'] = 0

    if progress_callback: progress_callback(0.5, "Харилцагч, давхардал...")
    
    # Ховор харилцагч
    try:
        cp_f = d['counterparty_name'].value_counts()
        d['cp_rare'] = (d['counterparty_name'].map(cp_f).fillna(0) <= 3).astype(int)
    except:
        d['cp_rare'] = 0

    # Ховор данс-харилцагч хос
    try:
        d['pair'] = d['account_code'] + '|' + d['counterparty_name']
        pf = d['pair'].value_counts()
        d['pair_rare'] = (d['pair'].map(pf).fillna(0) <= 2).astype(int)
    except:
        d['pair_rare'] = 0

    d['desc_empty'] = (d['transaction_description'].str.len() == 0).astype(int)

    # Давхардал
    try:
        d['dup_key'] = d['account_code'] + '|' + d['amount'].astype(str) + '|' + d['transaction_date']
        dk = d['dup_key'].value_counts()
        d['is_dup'] = (d['dup_key'].map(dk).fillna(1) > 1).astype(int)
    except:
        d['is_dup'] = 0

    if progress_callback: progress_callback(0.7, "Цаг хугацаа, тайлбар тулгалт...")
    
    # Цаг
    d['day'] = pd.to_numeric(d['transaction_date'].str[8:10], errors='coerce').fillna(15)
    d['month_num'] = pd.to_numeric(d['transaction_date'].str[5:7], errors='coerce').fillna(6)
    d['is_month_end'] = (d['day'] >= 28).astype(int)
    d['is_year_end'] = (d['month_num'] == 12).astype(int)

    # Тайлбар тулгалт
    d['desc_mismatch'] = 0
    d['name_no_overlap'] = 0
    d['dir_mismatch'] = 0
    try:
        stop_w = {'дансны','данс','нийт','бусад','зардал','орлого','төлбөр','хөрөнгө','тооцоо','бүртгэл','дүн','төгрөг','сая','мянга','журнал','гүйлгээ','баримт'}
        acct_words = {}
        for code in d['account_code'].unique():
            all_desc = ' '.join(d.loc[d['account_code'] == code, 'transaction_description'].str.lower())
            wc = Counter(re.findall(r'[а-яөүёА-ЯӨҮЁ\w]{3,}', all_desc))
            acct_words[code] = set(w for w, c in wc.items() if c >= 3 and len(w) >= 3)

        def _check_mismatch(code, tx_desc):
            tx = str(tx_desc).lower() if tx_desc else ''
            if not tx or code not in acct_words or not acct_words[code]: return 0
            tx_words = set(re.findall(r'[а-яөүёА-ЯӨҮЁ\w]{3,}', tx))
            return 0 if len(tx_words & acct_words[code]) > 0 else 1
        d['desc_mismatch'] = [_check_mismatch(c, t) for c, t in zip(d['account_code'], d['transaction_description'])]

        def _extract_kw(text):
            if not text: return set()
            return set(w for w in re.findall(r'[а-яөүёА-ЯӨҮЁ\w]{3,}', str(text).lower()) if w not in stop_w and len(w) >= 3)
        def _check_overlap(aname, tdesc):
            nk = _extract_kw(aname)
            dk2 = _extract_kw(tdesc)
            if not nk or not dk2: return 0
            return 0 if len(nk & dk2) > 0 else 1
        d['name_no_overlap'] = [_check_overlap(a, t) for a, t in zip(d['account_name'], d['transaction_description'])]
    except:
        pass

    # Чиглэл зөрчил
    try:
        af2 = d['account_code'].str[0]
        d.loc[(af2 == '1') & (d['credit_mnt'] > 0) & (d['debit_mnt'] == 0), 'dir_mismatch'] = 1
        d.loc[(af2 == '2') & (d['debit_mnt'] > 0) & (d['credit_mnt'] == 0), 'dir_mismatch'] = 1
        d.loc[(af2 == '5') & (d['debit_mnt'] > 0) & (d['credit_mnt'] == 0), 'dir_mismatch'] = 1
        d.loc[(af2.isin(['6','7','8'])) & (d['credit_mnt'] > 0) & (d['debit_mnt'] == 0), 'dir_mismatch'] = 1
    except:
        pass

    if progress_callback: progress_callback(0.9, "Audit Risk Rules...")
    
    # 🆕 v5.0 — Audit Risk Rules Engine
    d = AuditRiskEngine.apply_all_rules(d)

    if progress_callback: progress_callback(1.0, "Дууссан")
    
    return d


def run_txn_anomaly_v5(df, cont=0.05, use_advanced_ml=True):
    """v5.0 — Сайжруулсан аномали илрүүлэлт: IF + LOF + Autoencoder."""
    feats = ['log_amount','acct_cat_num','benford_dev','is_round','amt_zscore','cp_rare','pair_rare',
             'desc_empty','is_month_end','is_year_end','is_dup','is_debit','desc_mismatch',
             'name_no_overlap','dir_mismatch']
    
    # v5.0 нэмэлт features
    extra_feats = ['is_weekend','is_holiday','round_score','benford_acct_flag','sod_risk','near_materiality']
    for f in extra_feats:
        if f in df.columns:
            feats.append(f)
    
    for f in feats:
        if f not in df.columns:
            df[f] = 0
    
    X = df[feats].fillna(0).replace([np.inf, -np.inf], 0).astype(float)
    
    # Isolation Forest
    iso = IsolationForest(contamination=cont, random_state=42, n_estimators=200, n_jobs=-1)
    df['iso_anomaly'] = (iso.fit_predict(X) == -1).astype(int)
    df['iso_score'] = -iso.score_samples(X)
    
    if use_advanced_ml and len(df) > 100:
        # LOF
        try:
            lof_pred, lof_scores = MLEngine.lof_anomaly(X, contamination=cont)
            df['lof_anomaly'] = lof_pred
            df['lof_score'] = lof_scores
        except:
            df['lof_anomaly'] = 0
            df['lof_score'] = 0
        
        # Autoencoder
        try:
            ae_pred, ae_scores = MLEngine.autoencoder_anomaly(X, contamination=cont)
            df['ae_anomaly'] = ae_pred
            df['ae_score'] = ae_scores
        except:
            df['ae_anomaly'] = 0
            df['ae_score'] = 0
        
        # Weighted Ensemble (IF=0.4, LOF=0.3, AE=0.3)
        df['txn_anomaly'] = MLEngine.weighted_ensemble(
            [df['iso_anomaly'].values, df['lof_anomaly'].values, df['ae_anomaly'].values],
            weights=[0.4, 0.3, 0.3]
        )
    else:
        df['txn_anomaly'] = df['iso_anomaly']
    
    # Z-score flag
    try:
        z = np.abs(StandardScaler().fit_transform(X))
        df['txn_zscore_flag'] = (z.max(axis=1) > 2.5).astype(int)
    except:
        df['txn_zscore_flag'] = 0
    
    # Enhanced risk scoring
    df['txn_risk'] = (
        df['txn_anomaly'] * 3 + 
        df['txn_zscore_flag'] * 2 + 
        df['is_dup'] * 2 + 
        df['cp_rare'] +
        df['pair_rare'] + 
        (df['amt_zscore'].abs() > 3).astype(int) * 2 + 
        df['desc_empty'] +
        df['desc_mismatch'] * 2 + 
        df['name_no_overlap'] + 
        df['dir_mismatch'] * 2 +
        # v5.0 шинэ дүрмүүд
        df.get('is_weekend', pd.Series(0, index=df.index)).astype(int) * 2 +
        df.get('is_holiday', pd.Series(0, index=df.index)).astype(int) * 3 +
        df.get('benford_acct_flag', pd.Series(0, index=df.index)).astype(int) * 2 +
        df.get('sod_risk', pd.Series(0, index=df.index)).astype(int) * 2 +
        df.get('near_materiality', pd.Series(0, index=df.index)).astype(int) * 3
    )
    
    df['txn_risk_level'] = pd.cut(df['txn_risk'], bins=[-1, 3, 7, 12, 100],
        labels=['🟢 Бага','🟡 Дунд','🟠 Өндөр','🔴 Маш өндөр'])
    
    return df, feats


# ═══════════════════════════════════════
# ДАНСНЫ ТҮВШНИЙ ML (сайжруулсан)
# ═══════════════════════════════════════

def run_ml_v5(tb_all, cont, n_est, use_advanced=True):
    """v5.0 — Дансны түвшний ML шинжилгээ."""
    df = tb_all.copy()
    df['cat_code'] = df['account_code'].astype(str).str[:3]
    le = LabelEncoder()
    df['cat_num'] = le.fit_transform(df['cat_code'])
    df['log_turn_d'] = np.log1p(df['turnover_debit'].abs())
    df['log_turn_c'] = np.log1p(df['turnover_credit'].abs())
    df['log_close_d'] = np.log1p(df['closing_debit'].abs())
    df['log_close_c'] = np.log1p(df['closing_credit'].abs())
    df['turn_ratio'] = (df['turnover_debit'] / df['turnover_credit'].replace(0, np.nan)).fillna(0).replace([np.inf, -np.inf], 0)
    
    if 'net_change_signed' in df.columns:
        df['log_abs_change'] = np.log1p(df['net_change_signed'].abs())
    else:
        df['log_abs_change'] = np.log1p((df['closing_debit'] - df['opening_debit']).abs())
    
    feats = ['cat_num', 'log_turn_d', 'log_turn_c', 'log_close_d', 'log_close_c', 'turn_ratio', 'log_abs_change', 'year']
    X = df[feats].fillna(0).replace([np.inf, -np.inf], 0)
    
    # IF
    iso = IsolationForest(contamination=cont, random_state=42, n_estimators=200, n_jobs=-1)
    df['iso_anomaly'] = (iso.fit_predict(X) == -1).astype(int)
    
    # Z-score
    sc = StandardScaler()
    df['zscore_anomaly'] = (np.abs(sc.fit_transform(X)).max(axis=1) > 2.0).astype(int)
    
    # Turn ratio
    p95 = df['turn_ratio'].quantile(0.95)
    df['turn_anomaly'] = ((df['turn_ratio'] > p95) | (df['turn_ratio'] < -p95)).astype(int)
    
    # v5.0 — LOF, Autoencoder
    if use_advanced and len(df) > 50:
        try:
            lof_pred, _ = MLEngine.lof_anomaly(X, contamination=cont)
            df['lof_anomaly'] = lof_pred
        except:
            df['lof_anomaly'] = 0
        
        try:
            ae_pred, _ = MLEngine.autoencoder_anomaly(X, contamination=cont)
            df['ae_anomaly'] = ae_pred
        except:
            df['ae_anomaly'] = 0
    else:
        df['lof_anomaly'] = 0
        df['ae_anomaly'] = 0
    
    # Enhanced ensemble
    df['ensemble_anomaly'] = (
        (df['iso_anomaly'] == 1) | 
        ((df['zscore_anomaly'] == 1) & (df['turn_anomaly'] == 1)) |
        ((df['lof_anomaly'] == 1) & (df['ae_anomaly'] == 1))
    ).astype(int)
    
    y = df['ensemble_anomaly'].values
    cv = StratifiedKFold(n_splits=5, shuffle=True, random_state=42)
    
    models = {
        'Random Forest': RandomForestClassifier(n_estimators=n_est, max_depth=10, random_state=42, class_weight='balanced', n_jobs=-1),
        'Gradient Boosting': GradientBoostingClassifier(n_estimators=150, max_depth=5, learning_rate=0.1, random_state=42),
        'Logistic Regression': LogisticRegression(max_iter=1000, random_state=42, class_weight='balanced'),
    }
    
    res = {}
    for nm, mdl in models.items():
        yp = cross_val_predict(mdl, X, y, cv=cv, method='predict')
        ypr = cross_val_predict(mdl, X, y, cv=cv, method='predict_proba')[:, 1]
        res[nm] = {
            'pred': yp, 'prob': ypr,
            'precision': precision_score(y, yp, zero_division=0),
            'recall': recall_score(y, yp, zero_division=0),
            'f1': f1_score(y, yp, zero_division=0),
            'auc': roc_auc_score(y, ypr) if len(np.unique(y)) > 1 else 0,
        }
    
    best = max(res, key=lambda k: res[k]['f1'])
    rf = models['Random Forest']
    rf.fit(X, y)
    fi = pd.DataFrame({'feature': feats, 'importance': rf.feature_importances_}).sort_values('importance', ascending=False)
    
    # MUS simulation
    nt = len(df)
    ns = int(nt * 0.20)
    at = df['turnover_debit'].abs() + df['turnover_credit'].abs()
    wt = at / at.sum()
    wt = wt.fillna(1 / nt)
    np.random.seed(42)
    ms = np.zeros(nt, dtype=int)
    ms[np.random.choice(nt, size=ns, replace=False, p=wt.values)] = 1
    ym = (ms & y).astype(int)
    
    return df, X, y, feats, res, best, fi, ym


# ═══════════════════════════════════════
# ФАЙЛ ТАНИХ (сайжруулсан)
# ═══════════════════════════════════════

def detect_file_type(f):
    """v5.0 — Сайжруулсан файл таних: SAP, 1C, QuickBooks, Oracle."""
    name = f.name.lower()
    fname_orig = f.name
    year = get_year(f.name)

    if name.endswith('.csv') or name.endswith('.gz') or name.endswith('.csv.gz'):
        return 'ledger', year

    if not name.endswith('.xlsx') and not name.endswith('.xls'):
        return 'unknown', year

    name_check = fname_orig.lower().replace('_', ' ').replace('-', ' ')
    
    # ⚠️ TB шалгалтыг EDT-ийн ӨМНӨ хийх (TB GC, TB MB гэх мэт файлуудыг зөв таних)
    # TB_standardized
    if 'tb_standardized' in name_check or 'tb standardized' in name_check:
        return 'tb_std', year
    
    # ГҮЙЛГЭЭ_БАЛАНС / Trial Balance
    tb_keywords = ['гүйлгээ баланс', 'гүйлгээ_баланс', 'гуйлгээ баланс', 'trial balance',
                   'гүйлгэ баланс', 'гуйлгэ баланс', 'оборотно сальдовая', 'osv']
    for kw in tb_keywords:
        if kw in name_check:
            return 'raw_tb', year
    
    # TB + space/dot prefix (TB GC, TB MB, TB GCE гэх мэт)
    # "Journal, TB" нь EDT, гэвч "TB " эхэлсэн бол raw_tb
    tb_prefix = name_check.strip()
    if (tb_prefix.startswith('tb ') or tb_prefix.startswith('tb.') or 
        tb_prefix.startswith('tb_') or tb_prefix == 'tb'):
        return 'raw_tb', year
    
    # Part1
    if 'part1' in name_check or 'part 1' in name_check:
        return 'part1', year
    # Ledger
    if 'ledger' in name_check or 'prototype_ledger' in name_check:
        return 'ledger', year
    
    # EDT keywords (TB-ийн ДАРАА шалгана)
    edt_keywords = ['ерөнхий журнал', 'ерөнхий дэвтэр', 'едт', 'edt', 'general ledger', 'general journal',
                    'еренхий журнал', 'journal gc', 'journal entry', 'journal entries',
                    'gl detail', 'gl transactions', 'проводки', 'журнал проводок',
                    'sap fico', 'fb03', 'faglb03', 'journal']
    for kw in edt_keywords:
        if kw in name_check:
            return 'edt', year

    # Sheet structure detection
    import openpyxl
    try:
        raw = f.read(); f.seek(0)
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True)
        sheets = wb.sheetnames

        if '02_ACCOUNT_SUMMARY' in sheets:
            if '04_RISK_MATRIX' in sheets:
                wb.close(); return 'part1', year
            wb.close(); return 'tb_std', year

        if '04_RISK_MATRIX' in sheets:
            wb.close(); return 'part1', year

        ws = wb[sheets[0]]
        sample_rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            sample_rows.append(row)
            if i >= 300: break
        wb.close()

        for row in sample_rows:
            if row[0] is not None:
                s = str(row[0]).strip()
                if s.startswith('Данс:') or s.startswith('Компани:') or s.startswith('ЕРӨНХИЙ') or s.startswith('Журнал:'):
                    return 'edt', year
            for cell in row[:5]:
                if cell is not None and 'Данс:' in str(cell):
                    return 'edt', year

        for row in sample_rows:
            if len(row) >= 2 and row[1] is not None:
                code = str(row[1]).strip()
                if re.match(r'\d{3}-\d{2}-\d{2}-\d{3}', code):
                    return 'raw_tb', year

        # Check if it's a generic ERP GL export
        for i, row in enumerate(sample_rows[:30]):
            if row[0] is not None:
                headers = [str(c).strip() for c in row if c]
                erp = detect_erp_type(headers, sample_rows[i+1:i+20])
                if erp != 'generic':
                    return 'edt', year
                # Check for GL-like columns
                cm = auto_map_columns(headers)
                if ('debit_mnt' in cm or 'credit_mnt' in cm) and 'account_code' in cm:
                    return 'edt', year

        return 'unknown', year
    except Exception:
        f.seek(0)
        return 'unknown', year


def get_year(name):
    for y in range(2020, 2030):
        if str(y) in name:
            return y
    return 2025


def generate_part1(df_led, year):
    """Part1 файл үүсгэх."""
    df = df_led.copy()
    yr = str(year)
    df['debit_mnt'] = pd.to_numeric(df['debit_mnt'], errors='coerce').fillna(0)
    df['credit_mnt'] = pd.to_numeric(df['credit_mnt'], errors='coerce').fillna(0)
    df['balance_mnt'] = pd.to_numeric(df['balance_mnt'], errors='coerce').fillna(0)
    monthly = df.groupby(['month', 'account_code']).agg(
        total_debit_mnt=('debit_mnt', 'sum'), total_credit_mnt=('credit_mnt', 'sum'),
        ending_balance_mnt=('balance_mnt', 'last'), transaction_count=('debit_mnt', 'count')
    ).reset_index()
    monthly.insert(0, 'report_year', yr)
    anames = df.groupby('account_code')['account_name'].first()
    acct = df.groupby('account_code').agg(
        total_debit_mnt=('debit_mnt', 'sum'), total_credit_mnt=('credit_mnt', 'sum'),
        closing_balance_mnt=('balance_mnt', 'last')
    ).reset_index()
    acct['account_name'] = acct['account_code'].map(anames)
    acct.insert(0, 'report_year', yr)
    rm = df.groupby(['month', 'account_code', 'counterparty_name']).agg(
        transaction_count=('debit_mnt', 'count'), total_debit=('debit_mnt', 'sum'), total_credit=('credit_mnt', 'sum'),
    ).reset_index()
    rm['total_amount_mnt'] = rm['total_debit'].abs() + rm['total_credit'].abs()
    rm.insert(0, 'report_year', yr)
    p75a = rm['total_amount_mnt'].quantile(0.75)
    p75c = rm['transaction_count'].quantile(0.75)
    rm['risk_flag_large_txn'] = (rm['total_amount_mnt'] > p75a).astype(int)
    rm['risk_flag_high_frequency'] = (rm['transaction_count'] > p75c).astype(int)
    rm['risk_score'] = rm['risk_flag_large_txn'] + rm['risk_flag_high_frequency']
    rm['risk_level'] = pd.cut(rm['risk_score'], bins=[-0.1, 0.5, 1.5, 99], labels=['Бага', 'Дунд', 'Өндөр']).astype(str)
    rm['account_category'] = rm['account_code'].str[:1].map(
        {'1': 'Хөрөнгө', '2': 'Өр', '3': 'Эздийн өмч', '4': 'Зардал', '5': 'Орлого', '6': 'Орлого', '7': 'Зардал'}
    ).fillna('')
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as w:
        monthly.to_excel(w, sheet_name='02_MONTHLY_SUMMARY', index=False)
        acct.to_excel(w, sheet_name='03_ACCOUNT_SUMMARY', index=False)
        rm.to_excel(w, sheet_name='04_RISK_MATRIX', index=False)
    buf.seek(0)
    return buf, monthly, acct, rm, len(rm[rm['risk_score'] > 0])


def read_ledger(f):
    raw = f.read(); f.seek(0)
    if raw[:2] == b'\x1f\x8b':
        return pd.read_csv(io.StringIO(gzip.decompress(raw).decode('utf-8')), dtype={'account_code': str})
    encoding = detect_encoding(raw)
    return pd.read_csv(io.BytesIO(raw), dtype={'account_code': str}, encoding=encoding)


def load_tb(files):
    frames, stats = [], {}
    for f in files:
        year = get_year(f.name)
        df = pd.read_excel(f, sheet_name='02_ACCOUNT_SUMMARY')
        df['year'] = year
        for c in ['turnover_debit', 'turnover_credit', 'closing_debit', 'closing_credit', 'opening_debit', 'opening_credit']:
            if c in df.columns:
                df[c] = pd.to_numeric(df[c], errors='coerce').fillna(0)
        if 'net_change_signed' in df.columns:
            df['net_change_signed'] = pd.to_numeric(df['net_change_signed'], errors='coerce').fillna(0)
        stats[year] = {'accounts': len(df), 'turnover_d': df['turnover_debit'].sum(), 'turnover_c': df['turnover_credit'].sum()}
        frames.append(df)
    return pd.concat(frames, ignore_index=True), stats


def load_ledger_stats(files):
    stats, all_frames = {}, []
    for f in files:
        year = get_year(f.name)
        f.seek(0)
        df = read_ledger(f)
        for c in ['debit_mnt', 'credit_mnt']:
            df[c] = pd.to_numeric(df[c], errors='coerce').fillna(0)
        df['report_year'] = str(year)
        df = optimize_dtypes(df)  # v5.0 memory optimization
        mo = df.groupby('month').agg(rows=('debit_mnt', 'count'), debit=('debit_mnt', 'sum'), credit=('credit_mnt', 'sum'))
        stats[year] = {'rows': len(df), 'accounts': df['account_code'].nunique(), 'months': df['month'].nunique(), 'monthly': mo}
        all_frames.append(df)
    full_df = pd.concat(all_frames, ignore_index=True) if all_frames else pd.DataFrame()
    return stats, full_df


def load_part1(files):
    all_rm, all_mo = [], []
    for f in files:
        year = get_year(f.name)
        try:
            rm = pd.read_excel(f, sheet_name='04_RISK_MATRIX')
            rm['year'] = year; all_rm.append(rm)
        except: pass
        try:
            mo = pd.read_excel(f, sheet_name='02_MONTHLY_SUMMARY')
            mo['year'] = year; all_mo.append(mo)
        except: pass
    return (pd.concat(all_rm, ignore_index=True) if all_rm else pd.DataFrame(),
            pd.concat(all_mo, ignore_index=True) if all_mo else pd.DataFrame())


FILE_TYPE_LABELS = {
    'raw_tb': ('📗 ГҮЙЛГЭЭ_БАЛАНС', 'Гүйлгээ-балансын файл → TB болгон хөрвүүлнэ'),
    'edt': ('📘 ЕДТ / Ерөнхий журнал / GL Export', 'SAP, 1C, Oracle, QuickBooks, бүх ERP дэмжинэ'),
    'tb_std': ('📊 TB_standardized', 'Стандартчилсан гүйлгээ-баланс → бэлэн'),
    'ledger': ('📄 Ledger CSV/GZ', 'Ерөнхий дэвтрийн гүйлгээ → бэлэн'),
    'part1': ('📈 Part1', 'Эрсдэлийн нэгтгэл → бэлэн'),
    'unknown': ('❓ Тодорхойгүй', 'Файлын формат шалгана уу'),
}


# ═══════════════════════════════════════
# UI — SIDEBAR
# ═══════════════════════════════════════
with st.sidebar:
    st.header("📌 Цэс")
    page = st.radio("Алхам:", ["1️⃣ Өгөгдөл бэлтгэх", "2️⃣ Шинжилгээ"])
    
    st.markdown("---")
    st.markdown(f"""
    <div style="background:#f0f4ff; padding:10px; border-radius:8px; font-size:12px;">
    <b>v{APP_VERSION} Шинэ:</b><br>
    ✅ Universal Parser (SAP, 1C, Oracle)<br>
    ✅ ISA Risk Rules Engine<br>
    ✅ Autoencoder + LOF + DBSCAN<br>
    ✅ 3M+ мөрийн гүйцэтгэл<br>
    ✅ Нэмэлт 6+ audit rule<br>
    </div>
    """, unsafe_allow_html=True)

with st.expander("📖 Нэр томъёоны товч тайлбар", expanded=False):
    st.markdown("""
- **TB:** Гүйлгээ-баланс — эхний үлдэгдэл, эргэлт, эцсийн үлдэгдлийн нэгтгэл
- **Ledger / ЕДТ:** Ерөнхий дэвтрийн дэлгэрэнгүй гүйлгээ
- **Part1:** Сарын нэгтгэл + эрсдэлийн матриц
- **IF:** Isolation Forest — тусгаарлалтаар аномали олно
- **LOF:** Local Outlier Factor — ойрын цэгүүдтэй харьцуулж аномали олно 🆕
- **AE:** Autoencoder — мэдрэлийн сүлжээгээр сэргээлт алдаа хэмжинэ 🆕
- **XAI:** Тайлбарлагдах ХОУ — загварын шийдвэрийг тайлбарлана
- **MUS:** Мөнгөний нэгжид суурилсан уламжлалт түүвэрлэлт
- **ROC/AUC:** Загварын ялгах чадварын хэмжүүр
""")


# ═══════════════════════════════════════
# ХУУДАС 1: ӨГӨГДӨЛ БЭЛТГЭХ
# ═══════════════════════════════════════
if page.startswith("1"):
    st.header("1️⃣ Өгөгдөл бэлтгэх")
    st.markdown("""
    <div class="info-box">
        <b>📂 Universal Parser v5.0:</b> SAP FICO, Oracle GL, 1C, QuickBooks, ямар ч ERP-ийн файлыг автоматаар таниж хөрвүүлнэ.<br>
        <span style="color: #555; font-size: 13px;">
        Дэмжих формат: XLSX, CSV, GZ — хэдэн ч файл, ямар ч дараалал, ямар ч encoding
        </span>
    </div>
    """, unsafe_allow_html=True)

    uploaded = st.file_uploader("📎 Бүх файлуудаа энд оруулна уу", type=['xlsx','xls','csv','gz'], accept_multiple_files=True, key='smart_prep')

    if uploaded:
        detected = []
        for f in uploaded:
            ftype, year = detect_file_type(f)
            f.seek(0)
            detected.append({'file': f, 'type': ftype, 'year': year, 'name': f.name})

        st.markdown("### 🔍 Таних үр дүн")
        det_rows = []
        for d in detected:
            label, desc = FILE_TYPE_LABELS.get(d['type'], FILE_TYPE_LABELS['unknown'])
            det_rows.append({'Файл': d['name'], 'Төрөл': label, 'Он': d['year'], 'Тайлбар': desc})
        st.dataframe(pd.DataFrame(det_rows), use_container_width=True, hide_index=True)

        raw_tbs = [d for d in detected if d['type'] == 'raw_tb']
        edts = [d for d in detected if d['type'] == 'edt']
        unknowns = [d for d in detected if d['type'] == 'unknown']
        ready = [d for d in detected if d['type'] in ('tb_std', 'ledger', 'part1')]

        if unknowns:
            st.warning(f"⚠️ {len(unknowns)} файл танигдсангүй: {', '.join(u['name'] for u in unknowns)}")
        if ready:
            st.success(f"✅ Шинжилгээнд бэлэн: {len(ready)} файл → **2️⃣ Шинжилгээ** руу шилжээрэй")

        if raw_tbs or edts:
            if st.button("⚙️ Хөрвүүлэлт эхлүүлэх", type="primary", use_container_width=True):
                progress = st.progress(0)
                status = st.empty()
                total_files = len(raw_tbs) + len(edts)
                done = 0
                
                if raw_tbs:
                    if 'tb_res' not in st.session_state:
                        st.session_state.tb_res = {}
                    for d in raw_tbs:
                        status.text(f"📗 ГҮЙЛГЭЭ_БАЛАНС {d['year']} хөрвүүлж байна...")
                        d['file'].seek(0)
                        buf, tb_s = process_raw_tb(d['file'])
                        st.session_state.tb_res[d['year']] = {'buf': buf.getvalue(), 'tb': tb_s}
                        done += 1
                        progress.progress(done / total_files)
                        st.success(f"✅ TB {d['year']}: {len(tb_s):,} данс")
                
                if edts:
                    if 'led_res' not in st.session_state:
                        st.session_state.led_res = {}
                    edt_by_year = {}
                    for d in edts:
                        edt_by_year.setdefault(d['year'], []).append(d)
                    for yr in sorted(edt_by_year):
                        status.text(f"📘 ЕДТ {yr} хөрвүүлж байна ({len(edt_by_year[yr])} файл)...")
                        frames = []
                        failed_files = []
                        for d in edt_by_year[yr]:
                            d['file'].seek(0)
                            df_e, cnt_e = process_edt(d['file'], yr)
                            if cnt_e > 0:
                                frames.append(df_e)
                                st.success(f"  ✅ {d['name']}: {cnt_e:,} гүйлгээ")
                            else:
                                failed_files.append(d['name'])
                                st.warning(f"  ⚠️ {d['name']}: 0 гүйлгээ уншигдсан — ЕДТ формат биш байж магадгүй. Энэ файл ГҮЙЛГЭЭ_БАЛАНС (TB) формат байвал нэрийг нь \"TB_\" эхлүүлж өөрчилнө үү.")
                        if frames:
                            st.session_state.led_res[yr] = pd.concat(frames, ignore_index=True)
                            st.success(f"✅ ЕДТ {yr}: Нийт {len(st.session_state.led_res[yr]):,} гүйлгээ ({len(frames)} файлаас)")
                        elif failed_files:
                            st.error(f"❌ {yr} оны бүх ЕДТ файл(ууд) уншигдсангүй: {', '.join(failed_files)}")
                        done += 1
                        progress.progress(done / total_files)
                
                progress.progress(1.0)
                status.text("✅ Бүх хөрвүүлэлт дууссан!")

    # Download section
    if 'tb_res' in st.session_state and st.session_state.tb_res:
        st.markdown("---\n### 📥 TB файлууд")
        for yr in sorted(st.session_state.tb_res):
            d = st.session_state.tb_res[yr]
            st.download_button(f"📥 TB_standardized_{yr}.xlsx ({len(d['tb']):,} данс)", d['buf'], f"TB_standardized_{yr}1231.xlsx", key=f"dtb{yr}")

    if 'led_res' in st.session_state and st.session_state.led_res:
        st.markdown("---\n### 📥 Ledger + Part1")
        cols_out = ['report_year','account_code','account_name','transaction_no','transaction_date','journal_no','document_no','counterparty_name','counterparty_id','transaction_description','debit_mnt','credit_mnt','balance_mnt','month']
        for yr in sorted(st.session_state.led_res):
            dfy = st.session_state.led_res[yr]
            if dfy.empty: continue
            dfy['debit_mnt'] = pd.to_numeric(dfy['debit_mnt'], errors='coerce').fillna(0)
            dfy['credit_mnt'] = pd.to_numeric(dfy['credit_mnt'], errors='coerce').fillna(0)
            with st.expander(f"📅 {yr} — {len(dfy):,} гүйлгээ", expanded=True):
                p1_buf, p1_mo, p1_acct, p1_rm, n_risk = generate_part1(dfy, yr)
                c1x, c2x, c3x = st.columns(3)
                c1x.metric("Гүйлгээ", f"{len(dfy):,}")
                c2x.metric("Эрсдэлийн хос", f"{len(p1_rm):,}")
                c3x.metric("Эрсдэлтэй", f"{n_risk:,}")
                gz_bytes = gzip.compress(dfy[cols_out].to_csv(index=False).encode('utf-8'))
                st.download_button(f"📥 ledger_{yr}.csv.gz", gz_bytes, f"prototype_ledger_{yr}.csv.gz", key=f"dled{yr}")
                st.download_button(f"📥 part1_{yr}.xlsx", p1_buf.getvalue(), f"prototype_part1_{yr}.xlsx", key=f"dp1{yr}")


# ═══════════════════════════════════════
# ХУУДАС 2: ШИНЖИЛГЭЭ (using tab_descriptions from original)
# ═══════════════════════════════════════
elif page.startswith("2"):
    st.header("2️⃣ ХОУ Шинжилгээ v5.0")
    st.markdown("""
    <div style="background-color: #E8F5E9; padding: 15px; border-radius: 8px; border-left: 4px solid #2E7D32; margin-bottom: 15px;">
        <b>📂 Ямар ч файлаа нэг дор оруулаарай!</b> Universal Parser автоматаар таниж хөрвүүлнэ.<br>
        <span style="color: #555; font-size: 13px;">
        🆕 v5.0: Autoencoder + LOF + ISA Risk Rules + 3M+ мөрийн гүйцэтгэл
        </span>
    </div>
    """, unsafe_allow_html=True)

    all_files = st.file_uploader("📎 Бүх файл (ямар ч формат, хэдэн ч файл)", type=['xlsx','xls','csv','gz'], accept_multiple_files=True, key='smart_analysis')

    tb_files, led_files, p1_files = [], [], []

    if all_files:
        detected = []
        for f in all_files:
            ftype, year = detect_file_type(f)
            f.seek(0)
            detected.append({'file': f, 'type': ftype, 'year': year, 'name': f.name})

        det_rows = []
        for d in detected:
            label, desc = FILE_TYPE_LABELS.get(d['type'], FILE_TYPE_LABELS['unknown'])
            det_rows.append({'Файл': d['name'], 'Төрөл': label, 'Он': d['year'], 'Тайлбар': desc})
        st.dataframe(pd.DataFrame(det_rows), use_container_width=True, hide_index=True)

        raw_tbs = [d for d in detected if d['type'] == 'raw_tb']
        edts = [d for d in detected if d['type'] == 'edt']

        for d in detected:
            if d['type'] == 'tb_std': tb_files.append(d['file'])
            elif d['type'] == 'ledger': led_files.append(d['file'])
            elif d['type'] == 'part1': p1_files.append(d['file'])
            elif d['type'] == 'raw_tb':
                with st.spinner(f"📗 {d['name']} → TB..."):
                    d['file'].seek(0)
                    buf, _ = process_raw_tb(d['file'])
                    buf.seek(0)
                    tb_wrap = io.BytesIO(buf.getvalue())
                    tb_wrap.name = f"TB_standardized_{d['year']}1231.xlsx"
                    tb_files.append(tb_wrap)
                st.success(f"✅ {d['name']} → TB")
            elif d['type'] == 'edt':
                with st.spinner(f"📘 {d['name']} → Ledger + Part1..."):
                    d['file'].seek(0)
                    df_edt, cnt = process_edt(d['file'], d['year'])
                if cnt > 0 and not df_edt.empty:
                    cols_out = ['report_year','account_code','account_name','transaction_no','transaction_date',
                                'journal_no','document_no','counterparty_name','counterparty_id',
                                'transaction_description','debit_mnt','credit_mnt','balance_mnt','month']
                    df_edt['debit_mnt'] = pd.to_numeric(df_edt['debit_mnt'], errors='coerce').fillna(0)
                    df_edt['credit_mnt'] = pd.to_numeric(df_edt['credit_mnt'], errors='coerce').fillna(0)
                    csv_bytes = df_edt[cols_out].to_csv(index=False).encode('utf-8')
                    led_wrap = io.BytesIO(csv_bytes)
                    led_wrap.name = f"prototype_ledger_{d['year']}.csv"
                    led_files.append(led_wrap)
                    if 'edt_frames' not in st.session_state:
                        st.session_state['edt_frames'] = []
                    st.session_state['edt_frames'].append(df_edt)
                    p1_buf, _, _, _, _ = generate_part1(df_edt, d['year'])
                    p1_buf.seek(0)
                    p1_wrap = io.BytesIO(p1_buf.getvalue())
                    p1_wrap.name = f"prototype_part1_{d['year']}.xlsx"
                    p1_files.append(p1_wrap)
                    st.success(f"✅ {d['name']} → Ledger ({cnt:,}) + Part1")
                else:
                    st.warning(f"⚠️ {d['name']} — гүйлгээ уншигдсангүй")

    # Settings
    st.markdown('<div style="background:#F5F5F5; padding:12px; border-radius:8px; margin-bottom:10px;"><b>⚙️ Тохиргоо</b></div>', unsafe_allow_html=True)
    c1s, c2s, c3s = st.columns(3)
    with c1s:
        cont = st.slider("🎯 Contamination (аномалийн хувь)", 0.05, 0.20, 0.10, 0.01)
    with c2s:
        nest = st.slider("🌲 Модны тоо (RF)", 50, 500, 200, 50)
    with c3s:
        use_advanced = st.checkbox("🧠 Advanced ML (LOF + AE)", value=True, help="Autoencoder, LOF нэмэлт загварууд")

    has_any = tb_files or led_files
    if st.button("🚀 Шинжилгээ эхлүүлэх", type="primary", use_container_width=True) and has_any:
        start_time = time.time()
        
        df = pd.DataFrame(); X = np.array([]); y = np.array([]); feats = []
        res = {}; best = ''; fi = pd.DataFrame(); ym = np.array([])
        tb_st = {}; led_st = {}; ledger_full = pd.DataFrame()
        rm_all = pd.DataFrame(); mo_all = pd.DataFrame()

        if tb_files and led_files:
            with st.spinner("TB уншиж байна..."):
                tb_all, tb_st = load_tb(tb_files)
            with st.spinner("Ledger уншиж байна..."):
                led_st, ledger_full = load_ledger_stats(led_files)
            if p1_files:
                with st.spinner("Part1 уншиж байна..."):
                    rm_all, mo_all = load_part1(p1_files)
            with st.spinner("🤖 Дансны түвшний ML шинжилгээ (v5.0)..."):
                df, X, y, feats, res, best, fi, ym = run_ml_v5(tb_all, cont, nest, use_advanced)
        elif led_files:
            with st.spinner("Ledger уншиж байна..."):
                led_st, ledger_full = load_ledger_stats(led_files)
            if p1_files:
                with st.spinner("Part1 уншиж байна..."):
                    rm_all, mo_all = load_part1(p1_files)

        # Transaction analysis
        txn_result = pd.DataFrame()
        edt_frames = st.session_state.get('edt_frames', [])
        all_txn_frames = []
        if len(ledger_full) > 0:
            all_txn_frames.append(ledger_full)
        elif edt_frames:
            all_txn_frames.extend(edt_frames)

        if all_txn_frames:
            txn_combined = pd.concat(all_txn_frames, ignore_index=True)
            n_total = len(txn_combined)
            
            # v5.0 — Smart sampling for large datasets
            sample_n = min(n_total, SAMPLE_SIZE_DEFAULT)
            if n_total > sample_n:
                st.info(f"📊 {n_total:,} гүйлгээ → {sample_n:,} түүвэрлэлт (гүйцэтгэлийн хурдны тулд)")
            
            progress = st.progress(0)
            status = st.empty()
            
            try:
                txn_s = txn_combined.sample(n=sample_n, random_state=42) if n_total > sample_n else txn_combined.copy()
                
                status.text("🔧 Шинж чанар үүсгэж байна...")
                txn_s = engineer_txn_features_v5(txn_s, progress_callback=lambda p, t: (progress.progress(p * 0.7), status.text(t)))
                
                status.text("🤖 ML аномали илрүүлж байна...")
                progress.progress(0.75)
                txn_result, _ = run_txn_anomaly_v5(txn_s, cont, use_advanced)
                
                progress.progress(1.0)
                elapsed = time.time() - start_time
                status.text(f"✅ Дууссан! ({elapsed:.1f} секунд)")
            except Exception as e:
                st.warning(f"⚠️ Гүйлгээний шинжилгээ алдаа: {e}")

        # Store results
        for key, val in [('analysis_done', True), ('df', df), ('X', X), ('y', y), ('feats', feats),
                         ('res', res), ('best', best), ('fi', fi), ('ym', ym), ('tb_st', tb_st),
                         ('led_st', led_st), ('rm_all', rm_all), ('mo_all', mo_all), ('txn_result', txn_result)]:
            st.session_state[key] = val

    # Session state defaults
    for key, default in [('analysis_done', False), ('rm_all', pd.DataFrame()), ('mo_all', pd.DataFrame()), ('txn_result', pd.DataFrame())]:
        if key not in st.session_state:
            st.session_state[key] = default

    # ═══════════════════════════════════════
    # 4️⃣ DASHBOARD (сайжруулсан)
    # ═══════════════════════════════════════
    if st.session_state.get('analysis_done', False):
        df = st.session_state['df']
        X = st.session_state['X']
        y = st.session_state['y']
        feats = st.session_state['feats']
        res = st.session_state['res']
        best = st.session_state['best']
        fi = st.session_state['fi']
        ym = st.session_state['ym']
        tb_st = st.session_state['tb_st']
        led_st = st.session_state['led_st']
        rm_all = st.session_state['rm_all']
        mo_all = st.session_state['mo_all']
        txn_result = st.session_state.get('txn_result', pd.DataFrame())

        has_account = len(df) > 0 and len(res) > 0
        has_rm = len(rm_all) > 0
        has_mo = len(mo_all) > 0
        has_txn = len(txn_result) > 0
        n_led = sum(d['rows'] for d in led_st.values()) if led_st else (len(txn_result) if has_txn else 0)
        yrs = sorted(tb_st.keys()) if tb_st else []
        bp = res[best]['pred'] if has_account and best else np.array([])

        if has_account:
            st.success(f"✅ v5.0 | {len(df):,} данс, {n_led:,} гүйлгээ | IF + LOF + AE + Risk Rules")
        elif has_txn:
            st.success(f"✅ v5.0 | {len(txn_result):,} гүйлгээ шинжлэгдсэн")

        # Build tab list
        tab_names = []
        if has_account:
            tab_names.extend(["📊 Нэгтгэл", "🔍 Хэвийн бус данс", "⚖️ ХОУ ↔ MUS", "🧠 XAI", "📋 Жагсаалт"])
        if has_txn:
            tab_names.extend(["🔴 Гүйлгээний эрсдэл", "👤 Харилцагч"])
        if has_rm:
            tab_names.append("🎯 Эрсдэлийн матриц")
        if has_mo:
            tab_names.append("📈 Сарын хандлага")

        if not tab_names:
            st.info("👆 Файлуудаа оруулна уу.")
        else:
            all_tabs = st.tabs(tab_names)
            tab_idx = 0

            # TAB: НЭГТГЭЛ
            if has_account:
                with all_tabs[tab_idx]:
                    st.subheader("📊 Executive Summary")
                    m1, m2, m3, m4 = st.columns(4)
                    m1.metric("Данс", f"{len(df):,}")
                    m2.metric("Гүйлгээ", f"{n_led:,}")
                    m3.metric("Аномали", f"{df['ensemble_anomaly'].sum():,} ({df['ensemble_anomaly'].mean()*100:.1f}%)")
                    m4.metric("F1-score", f"{res[best]['f1']:.4f}")
                    
                    # v5.0 — ML method comparison
                    if has_account:
                        st.markdown("#### 🤖 Илрүүлэлтийн аргуудын нэгтгэл")
                        methods = {'IF': 'iso_anomaly', 'Z-score': 'zscore_anomaly', 'Turn ratio': 'turn_anomaly'}
                        if 'lof_anomaly' in df.columns:
                            methods['LOF 🆕'] = 'lof_anomaly'
                        if 'ae_anomaly' in df.columns:
                            methods['Autoencoder 🆕'] = 'ae_anomaly'
                        methods['ENSEMBLE'] = 'ensemble_anomaly'
                        
                        ad = []
                        for m, c in methods.items():
                            if c in df.columns:
                                row_d = {'Арга': m, 'Нийт': int(df[c].sum()), 'Хувь': f"{df[c].mean()*100:.1f}%"}
                                ad.append(row_d)
                        st.dataframe(pd.DataFrame(ad), use_container_width=True, hide_index=True)
                    
                    fg = make_subplots(rows=1, cols=3, subplot_titles=("Данс", "Эргэлт (T₮)", "ЕДТ мөр"))
                    cl3 = ['#2196F3', '#4CAF50', '#FF9800']
                    for i, yv in enumerate(yrs):
                        fg.add_trace(go.Bar(x=[str(yv)], y=[tb_st[yv]['accounts']], marker_color=cl3[i%3], showlegend=False), row=1, col=1)
                        fg.add_trace(go.Bar(x=[str(yv)], y=[tb_st[yv]['turnover_d']/1e9], marker_color=cl3[i%3], showlegend=False), row=1, col=2)
                        if yv in led_st:
                            fg.add_trace(go.Bar(x=[str(yv)], y=[led_st[yv]['rows']], marker_color=cl3[i%3], showlegend=False), row=1, col=3)
                    fg.update_layout(height=350)
                    st.plotly_chart(fg, use_container_width=True)
                tab_idx += 1

                # TAB: ХЭВИЙН БУС ДАНС
                with all_tabs[tab_idx]:
                    st.subheader("🔍 Хэвийн бус данс")
                    st.plotly_chart(px.scatter(df, x='log_turn_d', y='log_abs_change', 
                        color=df['ensemble_anomaly'].map({0:'Хэвийн',1:'Аномали'}),
                        facet_col='year', opacity=0.5, 
                        color_discrete_map={'Хэвийн':'#90caf9','Аномали':'#c62828'},
                        height=400, title="Аномали дансуудын тархалт"), use_container_width=True)
                tab_idx += 1

                # TAB: ХОУ ↔ MUS
                with all_tabs[tab_idx]:
                    st.subheader("⚖️ ХОУ загвар ↔ MUS 20%")
                    st.dataframe(pd.DataFrame([{'Загвар': n, 'Precision': f"{r['precision']:.4f}", 
                        'Recall': f"{r['recall']:.4f}", 'F1': f"{r['f1']:.4f}", 
                        'AUC': f"{r['auc']:.4f}"} for n, r in res.items()]), 
                        use_container_width=True, hide_index=True)
                    
                    fg2 = go.Figure()
                    for n, r in res.items():
                        fpr, tpr, _ = roc_curve(y, r['prob'])
                        fg2.add_trace(go.Scatter(x=fpr, y=tpr, name=f"{n} (AUC={r['auc']:.4f})"))
                    fg2.add_trace(go.Scatter(x=[0,1], y=[0,1], name='Random', line=dict(dash='dash', color='gray')))
                    fg2.update_layout(title='ROC Curve', height=400)
                    st.plotly_chart(fg2, use_container_width=True)
                    
                    # Detection Risk
                    dr = []
                    for yv in yrs:
                        mk = (df['year']==yv).values
                        yt = y[mk]; nt2 = yt.sum()
                        if nt2 > 0:
                            a2 = 1-(bp[mk]&yt).sum()/nt2
                            m2x = 1-(ym[mk]&yt).sum()/nt2
                        else: a2=0; m2x=0
                        dr.append({'Жил':yv, 'ХОУ':f"{a2:.4f}", 'MUS 20%':f"{m2x:.4f}", 'Сайжрал':f"{m2x-a2:.4f}"})
                    st.dataframe(pd.DataFrame(dr), use_container_width=True, hide_index=True)
                tab_idx += 1

                # TAB: XAI
                with all_tabs[tab_idx]:
                    st.subheader("🧠 Тайлбарлагдах ХОУ (XAI)")
                    st.plotly_chart(px.bar(fi, x='importance', y='feature', orientation='h', 
                        color='importance', color_continuous_scale='Blues',
                        title='Feature Importance').update_layout(height=400, yaxis={'categoryorder':'total ascending'}), 
                        use_container_width=True)
                tab_idx += 1

                # TAB: ЖАГСААЛТ
                with all_tabs[tab_idx]:
                    st.subheader("📋 Аномали дансуудын жагсаалт")
                    adf = df[df['ensemble_anomaly']==1][['year','account_code','account_name','turnover_debit','turnover_credit','turn_ratio','log_abs_change']].copy()
                    yf = st.selectbox("Жил", ['Бүгд']+[str(y2) for y2 in yrs])
                    if yf != 'Бүгд': adf = adf[adf['year']==int(yf)]
                    st.write(f"Нийт: {len(adf)}")
                    st.dataframe(adf, use_container_width=True, hide_index=True, height=500)
                    st.download_button("📥 CSV", adf.to_csv(index=False).encode('utf-8-sig'), "anomaly.csv")
                tab_idx += 1

            # TAB: ГҮЙЛГЭЭНИЙ ЭРСДЭЛ
            if has_txn:
                with all_tabs[tab_idx]:
                    st.subheader("🔴 Гүйлгээний эрсдэл (v5.0)")
                    
                    n_anom = txn_result['txn_anomaly'].sum()
                    c1,c2,c3,c4 = st.columns(4)
                    c1.metric("Шинжилсэн", f"{len(txn_result):,}")
                    c2.metric("Хэвийн бус", f"{n_anom:,}", delta=f"{n_anom/len(txn_result)*100:.1f}%", delta_color="inverse")
                    c3.metric("Тайлбар зөрчил", f"{txn_result['desc_mismatch'].sum():,}")
                    c4.metric("Чиглэл зөрсөн", f"{txn_result['dir_mismatch'].sum():,}")
                    
                    # v5.0 нэмэлт metrics
                    c5,c6,c7,c8 = st.columns(4)
                    c5.metric("Амралтын өдөр 🆕", f"{txn_result.get('is_weekend', pd.Series(0)).sum():,}")
                    c6.metric("Баярын өдөр 🆕", f"{txn_result.get('is_holiday', pd.Series(0)).sum():,}")
                    c7.metric("Бенфорд зөрчил 🆕", f"{txn_result.get('benford_acct_flag', pd.Series(0)).sum():,}")
                    c8.metric("SoD эрсдэл 🆕", f"{txn_result.get('sod_risk', pd.Series(0)).sum():,}")
                    
                    # Risk distribution
                    rl = txn_result['txn_risk_level'].value_counts().reindex(['🟢 Бага','🟡 Дунд','🟠 Өндөр','🔴 Маш өндөр']).fillna(0)
                    st.plotly_chart(px.bar(x=rl.index, y=rl.values, color=rl.index,
                        color_discrete_map={'🟢 Бага':'#4CAF50','🟡 Дунд':'#FFC107','🟠 Өндөр':'#FF9800','🔴 Маш өндөр':'#F44336'},
                        title="Эрсдэлийн тархалт").update_layout(height=300, showlegend=False), use_container_width=True)
                    
                    # Filtered list
                    fc1, fc2 = st.columns(2)
                    with fc1:
                        risk_f = st.selectbox("Эрсдэл:", ['Бүгд','🔴 Маш өндөр','🟠 Өндөр','🟡 Дунд'], key='txn_rf')
                    with fc2:
                        txn_years = sorted(txn_result['report_year'].dropna().unique().tolist()) if 'report_year' in txn_result.columns else []
                        year_f = st.selectbox("Он:", ['Бүгд']+[str(y) for y in txn_years], key='txn_yf')
                    
                    t_show = txn_result[txn_result['txn_anomaly']==1].copy() if risk_f=='Бүгд' else txn_result[txn_result['txn_risk_level']==risk_f].copy()
                    if year_f != 'Бүгд' and 'report_year' in t_show.columns:
                        t_show = t_show[t_show['report_year'].astype(str)==year_f]
                    
                    cols_show = ['txn_risk_level','txn_risk','report_year','account_code','account_name',
                                'counterparty_name','transaction_date','debit_mnt','credit_mnt',
                                'transaction_description','desc_mismatch','dir_mismatch','is_weekend','is_holiday']
                    t_disp = t_show[[c for c in cols_show if c in t_show.columns]].sort_values('txn_risk', ascending=False)
                    st.write(f"Нийт: **{len(t_disp):,}** гүйлгээ")
                    st.dataframe(t_disp, use_container_width=True, hide_index=True, height=500)
                    st.download_button("📥 CSV", t_disp.to_csv(index=False).encode('utf-8-sig'), "anomaly_txn.csv")
                tab_idx += 1

                # TAB: ХАРИЛЦАГЧ
                with all_tabs[tab_idx]:
                    st.subheader("👤 Харилцагчаар нэгтгэсэн")
                    cp_r = txn_result[txn_result['counterparty_name'].fillna('')!=''].groupby('counterparty_name').agg(
                        total=('amount','count'), anomaly=('txn_anomaly','sum'), amount=('amount','sum'),
                        accounts=('account_code','nunique')
                    ).reset_index()
                    cp_r['anomaly_pct'] = (cp_r['anomaly']/cp_r['total']*100).round(1)
                    cp_r = cp_r.sort_values('anomaly', ascending=False)
                    cp_r.columns = ['Харилцагч','Нийт','Хэвийн бус','Дүн','Данс','%']
                    st.dataframe(cp_r.head(50), use_container_width=True, hide_index=True)
                    
                    top20 = cp_r.head(20)
                    if len(top20)>0:
                        st.plotly_chart(px.bar(top20, x='Хэвийн бус', y='Харилцагч', orientation='h',
                            color='Данс', color_continuous_scale='Reds',
                            title='Топ 20 харилцагч').update_layout(height=500, yaxis={'categoryorder':'total ascending'}), 
                            use_container_width=True)
                tab_idx += 1

            # TAB: ЭРСДЭЛИЙН МАТРИЦ
            if has_rm:
                with all_tabs[tab_idx]:
                    st.subheader("🎯 Эрсдэлийн матриц")
                    rm_all['risk_score'] = pd.to_numeric(rm_all.get('risk_score', 0), errors='coerce').fillna(0)
                    fig_rm = go.Figure()
                    for yv in sorted(rm_all['year'].unique()):
                        rmy = rm_all[rm_all['year']==yv]
                        fig_rm.add_trace(go.Bar(x=['Нийт','Эрсдэлтэй'], y=[len(rmy), len(rmy[rmy['risk_score']>0])], name=str(yv)))
                    fig_rm.update_layout(barmode='group', height=350)
                    st.plotly_chart(fig_rm, use_container_width=True)
                tab_idx += 1

            # TAB: САРЫН ХАНДЛАГА
            if has_mo:
                with all_tabs[tab_idx]:
                    st.subheader("📈 Сарын хандлага")
                    mo_all['total_debit_mnt'] = pd.to_numeric(mo_all['total_debit_mnt'], errors='coerce').fillna(0)
                    mo_all['transaction_count'] = pd.to_numeric(mo_all['transaction_count'], errors='coerce').fillna(0)
                    mo_agg = mo_all.groupby('month').agg(debit=('total_debit_mnt','sum'), txn=('transaction_count','sum')).reset_index()
                    mo_agg['debit_T'] = mo_agg['debit']/1e9
                    fig_mo = make_subplots(rows=2, cols=1, subplot_titles=("Эргэлт (T₮)","Гүйлгээний тоо"))
                    fig_mo.add_trace(go.Scatter(x=mo_agg['month'], y=mo_agg['debit_T'], name='Дебит'), row=1, col=1)
                    fig_mo.add_trace(go.Bar(x=mo_agg['month'], y=mo_agg['txn'], name='Гүйлгээ'), row=2, col=1)
                    fig_mo.update_layout(height=500)
                    st.plotly_chart(fig_mo, use_container_width=True)

    if not st.session_state.get('analysis_done', False) and not (tb_files or led_files):
        st.info("👆 Файлуудаа оруулна уу. TB + Ledger = бүрэн шинжилгээ.")
